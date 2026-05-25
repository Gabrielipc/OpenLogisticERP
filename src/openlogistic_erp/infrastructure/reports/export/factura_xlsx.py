"""Template-based XLSX exporter for printable facturas."""

from __future__ import annotations

from collections.abc import Iterable, Sequence
from copy import copy
from decimal import Decimal, ROUND_HALF_UP
from importlib import resources
from pathlib import Path
from tempfile import NamedTemporaryFile
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

_TEMPLATE_PACKAGE = "openlogistic_erp.infrastructure.templates.factura"
_TEMPLATE_NAME = "formato_factura.xlsx"
_TEMPLATE_SHEET = "formato"
_DETAIL_START_ROW = 7
_DETAIL_END_ROW = 16
_DETAIL_CAPACITY = _DETAIL_END_ROW - _DETAIL_START_ROW + 1
_DETAIL_STYLE_ROW = 8
_TOTAL_ROW = 17
_FINAL_TOTAL_ROW = 25


class FacturaExcelExporter:
    """Export facturas to a workbook with one printable sheet per factura."""

    def export(self, facturas: Sequence[Any] | Iterable[Any], target_path: str | Path) -> None:
        invoice_rows = list(facturas)
        if not invoice_rows:
            raise ValueError("Se requiere al menos una factura para exportar.")

        path = Path(target_path)
        path.parent.mkdir(parents=True, exist_ok=True)

        workbook = _load_template_workbook()
        template = workbook[_TEMPLATE_SHEET]
        used_titles: set[str] = set()
        for factura in invoice_rows:
            sheet = workbook.copy_worksheet(template)
            sheet.title = _unique_sheet_title(_sheet_base_title(factura), used_titles)
            _write_factura_sheet(sheet, factura)

        workbook.remove(template)
        workbook.save(path)


def _load_template_workbook():
    template = resources.files(_TEMPLATE_PACKAGE).joinpath(_TEMPLATE_NAME)
    with resources.as_file(template) as template_path:
        return load_workbook(template_path)


def _write_factura_sheet(sheet: Worksheet, factura: Any) -> None:
    detalles = _detail_rows(factura)
    extra_rows = max(0, len(detalles) - _DETAIL_CAPACITY)
    if extra_rows:
        sheet.insert_rows(_TOTAL_ROW, amount=extra_rows)

    total_row = _TOTAL_ROW + extra_rows
    final_total_row = _FINAL_TOTAL_ROW + extra_rows
    last_detail_row = _DETAIL_START_ROW + max(len(detalles), _DETAIL_CAPACITY) - 1

    _unmerge_detail_cells(sheet, _DETAIL_START_ROW, last_detail_row)
    _copy_detail_row_style(sheet, _DETAIL_STYLE_ROW, _DETAIL_START_ROW, last_detail_row)
    _clear_detail_rows(sheet, _DETAIL_START_ROW, last_detail_row)

    cliente = getattr(factura, "cliente", None)
    sheet["A2"] = str(getattr(cliente, "nombre", "") or "")
    sheet["D2"] = str(getattr(cliente, "ruc", "") or "")
    sheet["C3"] = getattr(factura, "fecha_emision", None)

    for offset, detail in enumerate(detalles):
        row = _DETAIL_START_ROW + offset
        sheet.cell(row=row, column=2).value = detail["referencia"] or None
        sheet.cell(row=row, column=3).value = detail["descripcion"] or None
        sheet.cell(row=row, column=4).value = detail["total"]

    sheet.cell(row=total_row, column=2).value = _amount_words(_sum_detail_totals(detalles), _currency_value(factura))
    sheet.cell(row=total_row, column=4).value = f"=SUM(D{_DETAIL_START_ROW}:D{last_detail_row})"
    sheet.cell(row=final_total_row, column=4).value = f"=D{total_row}"


def _unmerge_detail_cells(sheet: Worksheet, start_row: int, end_row: int) -> None:
    for merged_range in list(sheet.merged_cells.ranges):
        if start_row <= merged_range.min_row <= end_row or start_row <= merged_range.max_row <= end_row:
            sheet.unmerge_cells(str(merged_range))


def _copy_detail_row_style(sheet: Worksheet, source_row: int, start_row: int, end_row: int) -> None:
    source_height = sheet.row_dimensions[source_row].height
    for row in range(start_row, end_row + 1):
        sheet.row_dimensions[row].height = source_height
        for column in range(2, 5):
            source_cell = sheet.cell(row=source_row, column=column)
            target_cell = sheet.cell(row=row, column=column)
            if target_cell is source_cell:
                continue
            target_cell._style = copy(source_cell._style)
            if source_cell.has_style:
                target_cell.font = copy(source_cell.font)
                target_cell.fill = copy(source_cell.fill)
                target_cell.border = copy(source_cell.border)
                target_cell.alignment = copy(source_cell.alignment)
                target_cell.number_format = source_cell.number_format
                target_cell.protection = copy(source_cell.protection)


def _clear_detail_rows(sheet: Worksheet, start_row: int, end_row: int) -> None:
    for row in range(start_row, end_row + 1):
        for column in range(2, 5):
            sheet.cell(row=row, column=column).value = None


def _detail_rows(factura: Any) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for detail in getattr(factura, "detalles", []) or []:
        viaje = getattr(detail, "viaje", None)
        gasto = getattr(detail, "gasto", None)
        rows.append(
            {
                "referencia": str(getattr(viaje, "referencia", "") or "") if viaje is not None else "",
                "descripcion": _detail_description(detail, viaje, gasto),
                "total": _money_value(getattr(detail, "costo", None)),
            }
        )
    return rows


def _detail_description(detail: Any, viaje: Any, gasto: Any) -> str:
    descripcion = ""
    if viaje is not None:
        descripcion = str(getattr(viaje, "descripcion", "") or "")
    elif gasto is not None:
        descripcion = str(getattr(gasto, "descripcion", "") or "")
    if descripcion:
        return descripcion
    return str(getattr(detail, "descripcion", "") or "")


def _money_value(value: Any) -> Decimal:
    if value is None or value == "":
        return Decimal("0.00")
    return Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def _sum_detail_totals(detalles: list[dict[str, Any]]) -> Decimal:
    return sum((_money_value(detail["total"]) for detail in detalles), Decimal("0.00"))


def _currency_value(factura: Any) -> str:
    moneda = getattr(factura, "moneda", "")
    return str(getattr(moneda, "value", moneda) or "").upper()


def _amount_words(amount: Decimal, currency: str) -> str:
    integer_amount = int(amount.quantize(Decimal("1"), rounding=ROUND_HALF_UP))
    currency_label = "DOLARES" if currency == "USD" else "CORDOBAS"
    return f"{_number_to_spanish(integer_amount)} {currency_label}.".upper()


_UNITS = {
    0: "cero",
    1: "uno",
    2: "dos",
    3: "tres",
    4: "cuatro",
    5: "cinco",
    6: "seis",
    7: "siete",
    8: "ocho",
    9: "nueve",
    10: "diez",
    11: "once",
    12: "doce",
    13: "trece",
    14: "catorce",
    15: "quince",
    16: "dieciseis",
    17: "diecisiete",
    18: "dieciocho",
    19: "diecinueve",
    20: "veinte",
    30: "treinta",
    40: "cuarenta",
    50: "cincuenta",
    60: "sesenta",
    70: "setenta",
    80: "ochenta",
    90: "noventa",
}

_HUNDREDS = {
    100: "cien",
    200: "doscientos",
    300: "trescientos",
    400: "cuatrocientos",
    500: "quinientos",
    600: "seiscientos",
    700: "setecientos",
    800: "ochocientos",
    900: "novecientos",
}


def _number_to_spanish(value: int) -> str:
    if value < 0:
        return f"menos {_number_to_spanish(abs(value))}"
    if value < 21:
        return _UNITS[value]
    if value < 30:
        return f"veinti{_UNITS[value - 20]}"
    if value < 100:
        tens = (value // 10) * 10
        units = value % 10
        return _UNITS[tens] if units == 0 else f"{_UNITS[tens]} y {_UNITS[units]}"
    if value < 1000:
        hundreds = (value // 100) * 100
        rest = value % 100
        if value == 100:
            return _HUNDREDS[100]
        prefix = "ciento" if hundreds == 100 else _HUNDREDS[hundreds]
        return prefix if rest == 0 else f"{prefix} {_number_to_spanish(rest)}"
    if value < 1_000_000:
        thousands = value // 1000
        rest = value % 1000
        prefix = "mil" if thousands == 1 else f"{_number_to_spanish(thousands)} mil"
        return prefix if rest == 0 else f"{prefix} {_number_to_spanish(rest)}"
    millions = value // 1_000_000
    rest = value % 1_000_000
    prefix = "un millon" if millions == 1 else f"{_number_to_spanish(millions)} millones"
    return prefix if rest == 0 else f"{prefix} {_number_to_spanish(rest)}"


def _sheet_base_title(factura: Any) -> str:
    return str(getattr(factura, "numero_factura", "") or getattr(factura, "id", "") or "Factura")


def _unique_sheet_title(raw_title: str, used_titles: set[str]) -> str:
    invalid_chars = set('[]:*?/\\')
    base = "".join("_" if char in invalid_chars else char for char in raw_title).strip()[:31] or "Factura"
    title = base
    counter = 2
    while title in used_titles:
        suffix = f" {counter}"
        title = f"{base[: 31 - len(suffix)]}{suffix}"
        counter += 1
    used_titles.add(title)
    return title

