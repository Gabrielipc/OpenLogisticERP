from __future__ import annotations

from datetime import datetime
from decimal import Decimal
from types import SimpleNamespace

from openpyxl import load_workbook

from openlogistic_erp.infrastructure.reports.export import FacturaExcelExporter
from tests.builders.modelo_seed import build_factura_payload, create_cliente


def _factura(
    numero: str = "FAC-001",
    *,
    cliente_nombre: str = "Cliente Demo",
    cliente_ruc: str = "J0310000410607",
    detalles: list[SimpleNamespace] | None = None,
):
    return SimpleNamespace(
        id=1,
        numero_factura=numero,
        fecha_emision=datetime(2026, 5, 20, 9, 30),
        moneda="USD",
        cliente=SimpleNamespace(nombre=cliente_nombre, ruc=cliente_ruc),
        detalles=detalles
        if detalles is not None
        else [
            SimpleNamespace(
                tipo="Viaje",
                costo=Decimal("1020.00"),
                viaje=SimpleNamespace(referencia="2060", descripcion="Flete terrestre El Salvador - Managua"),
                gasto=None,
            ),
            SimpleNamespace(
                tipo="Gasto",
                costo=Decimal("80.00"),
                viaje=None,
                gasto=SimpleNamespace(descripcion="Un dia de estadia"),
            ),
        ],
    )


def test_factura_excel_exporter_writes_invoice_header_details_and_totals(tmp_path):
    target = tmp_path / "facturas.xlsx"

    FacturaExcelExporter().export([_factura()], target)

    workbook = load_workbook(target, data_only=False)
    assert workbook.sheetnames == ["FAC-001"]
    sheet = workbook["FAC-001"]
    assert sheet["A2"].value == "Cliente Demo"
    assert sheet["D2"].value == "J0310000410607"
    assert sheet["C3"].value == datetime(2026, 5, 20, 9, 30)
    assert sheet["B7"].value == "2060"
    assert sheet["C7"].value == "Flete terrestre El Salvador - Managua"
    assert sheet["D7"].value == Decimal("1020.00")
    assert sheet["B8"].value is None
    assert sheet["C8"].value == "Un dia de estadia"
    assert sheet["D8"].value == Decimal("80.00")
    assert sheet["D17"].value == "=SUM(D7:D16)"
    assert sheet["D25"].value == "=D17"


def test_factura_excel_exporter_writes_one_sheet_per_invoice(tmp_path):
    target = tmp_path / "facturas.xlsx"

    FacturaExcelExporter().export(
        [
            _factura("FAC-001", cliente_nombre="Cliente Uno"),
            _factura("FAC-002", cliente_nombre="Cliente Dos"),
        ],
        target,
    )

    workbook = load_workbook(target, data_only=False)
    assert workbook.sheetnames == ["FAC-001", "FAC-002"]
    assert workbook["FAC-001"]["A2"].value == "Cliente Uno"
    assert workbook["FAC-002"]["A2"].value == "Cliente Dos"


def test_factura_excel_exporter_inserts_extra_detail_rows_with_template_style(tmp_path):
    target = tmp_path / "facturas.xlsx"
    details = [
        SimpleNamespace(
            tipo="Viaje",
            costo=Decimal(str(index)),
            viaje=SimpleNamespace(referencia=f"REF-{index:02d}", descripcion=f"Detalle {index:02d}"),
            gasto=None,
        )
        for index in range(1, 13)
    ]

    FacturaExcelExporter().export([_factura(detalles=details)], target)

    workbook = load_workbook(target, data_only=False)
    sheet = workbook["FAC-001"]
    assert sheet["B18"].value == "REF-12"
    assert sheet["C18"].value == "Detalle 12"
    assert sheet["D18"].value == Decimal("12")
    assert sheet.row_dimensions[18].height == sheet.row_dimensions[8].height
    assert sheet["B18"].style_id == sheet["B8"].style_id
    assert sheet["C18"].style_id == sheet["C8"].style_id
    assert sheet["D18"].style_id == sheet["D8"].style_id
    assert sheet["D19"].value == "=SUM(D7:D18)"
    assert sheet["D27"].value == "=D19"
    assert sheet["A22"].value == "Favor elaborar CK o transferencia a nombre de:"


def test_factura_workflow_service_exports_selected_invoices_to_excel(
    modelo_workflow,
    session_factory,
    tmp_path,
):
    with session_factory() as session:
        cliente = create_cliente(session, nombre="Cliente Servicio", ruc="J-SERVICE-001")
        session.commit()

    payload = build_factura_payload(cliente.id, [])
    payload["factura"]["numero_factura"] = "FAC-SERVICE-001"
    factura = modelo_workflow.factura.create(payload)
    target = tmp_path / "factura_servicio.xlsx"

    returned_path = modelo_workflow.factura.export_excel([int(factura["id"])], target)

    workbook = load_workbook(returned_path, data_only=False)
    assert workbook.sheetnames == ["FAC-SERVICE-001"]
    sheet = workbook["FAC-SERVICE-001"]
    assert sheet["A2"].value == "Cliente Servicio"
    assert sheet["D2"].value == "J-SERVICE-001"
    assert sheet["D17"].value == "=SUM(D7:D16)"
