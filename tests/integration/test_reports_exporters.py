from __future__ import annotations

from datetime import date, datetime

from openpyxl import load_workbook
import pytest

from openlogistic_erp.application.reports import ReportExportService
from openlogistic_erp.application.reports.errors import ReportExportError
from openlogistic_erp.application.reports.layouts import (
    build_default_report_layouts,
    layout_for,
)
from openlogistic_erp.domain.reports import ReportColumn, ReportFormat, ReportPayload, ReportTable
from openlogistic_erp.infrastructure.reports.export import pdf as pdf_export
from openlogistic_erp.infrastructure.reports.export import PdfReportExporter, XlsxReportExporter
from openlogistic_erp.infrastructure.reports.export.formatting import format_cell


def sample_payload() -> ReportPayload:
    return ReportPayload(
        title="Reporte de prueba",
        generated_at=datetime(2026, 5, 6, 9, 45),
        message="Datos de muestra",
        tables=(
            ReportTable(
                key="detalle",
                title="Detalle",
                currency_field="moneda",
                columns=(
                    ReportColumn("fecha", "Fecha", "date"),
                    ReportColumn("cliente", "Cliente"),
                    ReportColumn("moneda", "Moneda"),
                    ReportColumn("monto", "Monto", "currency", decimals=2),
                    ReportColumn("margen", "Margen", "percent", decimals=1),
                ),
                rows=(
                    {
                        "fecha": date(2026, 5, 1),
                        "cliente": "Acme",
                        "moneda": "USD",
                        "monto": 1234.5,
                        "margen": 0.253,
                    },
                    {
                        "fecha": date(2026, 5, 2),
                        "cliente": "Beta",
                        "moneda": "NIO",
                        "monto": 800,
                        "margen": 0.1,
                    },
                ),
            ),
        ),
        currencies=({"key": "USD", "label": "US Dollar"}, {"key": "NIO", "label": "Cordoba"}),
    )


def test_format_cell_handles_numbers_dates_and_empty_values():
    assert format_cell(ReportColumn("n", "N", "int"), 1200) == "1,200"
    assert format_cell(ReportColumn("n", "N", "float", decimals=1), 12.34) == "12.3"
    assert format_cell(ReportColumn("p", "P", "percent", decimals=1), 0.253) == "25.3%"
    assert (
        format_cell(ReportColumn("d", "D", "datetime"), datetime(2026, 5, 6, 9, 45))
        == "2026-05-06 09:45"
    )
    assert format_cell(ReportColumn("x", "X"), None) == ""


def test_format_cell_handles_date_currency_decimals_and_invalid_numeric_fallback():
    assert format_cell(ReportColumn("d", "D", "date"), date(2026, 5, 6)) == "2026-05-06"
    assert format_cell(ReportColumn("c", "C", "currency", decimals=2), 1234.5) == "1,234.50"
    assert format_cell(ReportColumn("c", "C", "currency", decimals=2), 1234.5, currency="USD") == "$ 1,234.50"
    assert format_cell(ReportColumn("c", "C", "currency", decimals=2), 800, currency="NIO") == "C$ 800.00"
    assert format_cell(ReportColumn("n", "N", "int"), "sin dato") == "sin dato"


def test_default_report_layouts_include_expected_kpis_and_sheet_names():
    layouts = build_default_report_layouts()

    assert list(layouts) == [
        "viajes_por_conductor",
        "cuentas_por_cobrar_aging",
        "facturacion_por_cliente",
        "estado_cuenta_cliente",
    ]
    assert tuple(kpi.key for kpi in layouts["viajes_por_conductor"].kpis) == (
        "conductores_activos",
        "viajes_totales",
        "dias_ocupados",
        "promedio_dias_viaje",
    )
    assert tuple(kpi.key for kpi in layouts["estado_cuenta_cliente"].kpis) == (
        "total_facturado",
        "total_cobrado",
        "saldo_pendiente",
    )
    assert layouts["facturacion_por_cliente"].sheet_names["detalle"] == "Detalle de facturacion"


def test_layout_for_returns_registered_layout_or_none():
    assert layout_for("cuentas_por_cobrar_aging").title == "Cuentas por cobrar aging"
    assert layout_for("desconocido") is None


def test_xlsx_report_exporter_writes_non_empty_file(tmp_path):
    target = tmp_path / "reporte.xlsx"

    XlsxReportExporter().export(sample_payload(), target, currency_key="USD")

    assert target.is_file()
    assert target.stat().st_size > 0


def test_xlsx_report_exporter_honors_column_decimal_formats(tmp_path):
    target = tmp_path / "reporte.xlsx"
    payload = ReportPayload(
        title="Reporte de formatos",
        generated_at=datetime(2026, 5, 6, 9, 45),
        tables=(
            ReportTable(
                key="detalle",
                title="Detalle",
                columns=(
                    ReportColumn("valor", "Valor", "float", decimals=3),
                    ReportColumn("margen", "Margen", "percent", decimals=1),
                ),
                rows=({"valor": 12.3456, "margen": 0.253},),
            ),
        ),
    )

    XlsxReportExporter().export(payload, target)

    workbook = load_workbook(target, data_only=False)
    sheet = workbook["Detalle"]
    assert sheet["A2"].number_format == "#,##0.000"
    assert sheet["B2"].number_format == "0.0%"


def test_xlsx_report_exporter_writes_formula_like_text_as_strings(tmp_path):
    target = tmp_path / "reporte.xlsx"
    payload = ReportPayload(
        title="Reporte de texto",
        generated_at=datetime(2026, 5, 6, 9, 45),
        message="=SUM(1,2)",
        tables=(
            ReportTable(
                key="detalle",
                title="Detalle",
                columns=(ReportColumn("cliente", "Cliente"),),
                rows=({"cliente": "=SUM(1,2)"},),
            ),
        ),
    )

    XlsxReportExporter().export(payload, target)

    workbook = load_workbook(target, data_only=False)
    cover_cell = workbook["Portada"]["B5"]
    table_cell = workbook["Detalle"]["A2"]
    assert cover_cell.value == "=SUM(1,2)"
    assert cover_cell.data_type == "s"
    assert table_cell.value == "=SUM(1,2)"
    assert table_cell.data_type == "s"


def test_pdf_report_exporter_writes_non_empty_file(tmp_path, qapp):
    del qapp
    target = tmp_path / "reporte.pdf"

    PdfReportExporter().export(sample_payload(), target, currency_key="USD")

    assert target.is_file()
    assert target.stat().st_size > 0


def test_pdf_report_exporter_raises_when_print_does_not_create_file(tmp_path, monkeypatch):
    target = tmp_path / "reporte.pdf"

    class StubDocument:
        def setHtml(self, html: str) -> None:
            del html

        def print_(self, printer) -> None:
            del printer

    class StubPrinter:
        class PrinterMode:
            HighResolution = object()

        class OutputFormat:
            PdfFormat = object()

        def __init__(self, mode) -> None:
            del mode

        def setOutputFormat(self, output_format) -> None:
            del output_format

        def setOutputFileName(self, output_file_name: str) -> None:
            del output_file_name

    monkeypatch.setattr(pdf_export, "QTextDocument", StubDocument)
    monkeypatch.setattr(pdf_export, "QPrinter", StubPrinter)

    with pytest.raises(RuntimeError, match="PDF export failed.*empty"):
        PdfReportExporter().export(sample_payload(), target)


def test_report_export_service_dispatches_by_format_and_string_format(tmp_path):
    calls: list[tuple[ReportPayload, str, str]] = []

    class StubExporter:
        def export(self, payload: ReportPayload, target_path, currency_key: str = "") -> None:
            calls.append((payload, str(target_path), currency_key))

    target = tmp_path / "reporte.xlsx"
    payload = sample_payload()
    service = ReportExportService({ReportFormat.XLSX: StubExporter()})

    returned_path = service.export(payload, target, "xlsx", currency_key="USD")

    assert returned_path == str(target)
    assert calls == [(payload, str(target), "USD")]


def test_report_export_service_rejects_empty_target_path_and_unsupported_format(tmp_path):
    service = ReportExportService({})

    with pytest.raises(ReportExportError, match="target path"):
        service.export(sample_payload(), "", ReportFormat.XLSX)

    with pytest.raises(ReportExportError, match="Unsupported"):
        service.export(sample_payload(), tmp_path / "reporte.csv", ReportFormat.CSV)
